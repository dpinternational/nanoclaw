#!/usr/bin/env python3
"""Parse the DAILY tab from Production Reporting xlsx."""
import openpyxl, json
from collections import defaultdict

wb = openpyxl.load_workbook('/home/david/nanoclaw/data/production-reporting.xlsx', data_only=True)
ws = wb['DAILY']
print(f'DAILY tab: {ws.max_row} rows x {ws.max_column} cols')

dates = []
for col in range(3, min(ws.max_column+1, 430)):
    val = ws.cell(1, col).value
    dates.append(str(val)[:10] if val else None)

print(f'Date columns: {len(dates)}')
last_dates = [d for d in dates if d and d != 'None']
print(f'Date range: {last_dates[0] if last_dates else "?"} to {last_dates[-1] if last_dates else "?"}')

agent_totals = {}
for row_idx in range(2, min(ws.max_row+1, 420)):
    name = ws.cell(row_idx, 1).value
    if not name: continue
    name = str(name).strip()
    monthly = defaultdict(float)
    for col_idx in range(3, min(ws.max_column+1, 430)):
        val = ws.cell(row_idx, col_idx).value
        if val and col_idx-3 < len(dates) and dates[col_idx-3] and dates[col_idx-3] != 'None':
            try:
                amount = float(val)
                if amount > 0:
                    month = dates[col_idx-3][:7]
                    monthly[month] += amount
            except (ValueError, TypeError): pass
    if monthly:
        agent_totals[name] = dict(monthly)

print(f'\nAgents with production: {len(agent_totals)}')

sorted_agents = sorted(agent_totals.items(), key=lambda x: -sum(x[1].values()))
print(f'\nTop 20 producers:')
for name, monthly in sorted_agents[:20]:
    total = sum(monthly.values())
    months = len(monthly)
    print(f'  {name:30s} ${total:>12,.2f}  ({months} months)')

monthly_totals = defaultdict(float)
monthly_agents = defaultdict(set)
for name, monthly in agent_totals.items():
    for month, amount in monthly.items():
        monthly_totals[month] += amount
        monthly_agents[month].add(name)

print(f'\nMonthly totals:')
for month in sorted(monthly_totals.keys()):
    print(f'  {month}: ${monthly_totals[month]:>12,.2f}  ({len(monthly_agents[month])} agents)')

json.dump(agent_totals, open('/home/david/nanoclaw/data/daily-production-by-agent.json', 'w'))
print(f'\nSaved {len(agent_totals)} agents')
