#!/usr/bin/env python3
"""
Gina's Production Report Sync — pulls data from the Production Reporting xlsx
on Google Drive and loads it into the TPG Intel DB.

Syncs:
  1. Team production by manager (monthly)
  2. Personal production by agent (monthly)  
  3. Monthly top 10 rankings
  4. Screen names → real names mapping
  
Run on server: python3 scripts/gina-production-sync.py
"""

import json
import os
import sqlite3
import subprocess
import sys
from datetime import datetime

NANOCLAW = "/home/david/nanoclaw"
INTEL_DB = os.path.join(NANOCLAW, "store", "tpg-intel.db")
XLSX_PATH = os.path.join(NANOCLAW, "data", "production-reporting.xlsx")
DRIVE_FILE_ID = "1nW9OSnOiscmQ6HOgmZoV0cdgB1TvyKQO"


def download_latest():
    """Download the latest xlsx from Google Drive."""
    script = f"""
const {{google}} = require('googleapis');
const fs = require('fs'), path = require('path');
const credDir = '/home/david/.gmail-mcp';
const keys = JSON.parse(fs.readFileSync(path.join(credDir, 'gcp-oauth.keys.json'), 'utf-8'));
const tokens = JSON.parse(fs.readFileSync(path.join(credDir, 'credentials.json'), 'utf-8'));
const config = keys.installed || keys.web || keys;
const auth = new google.auth.OAuth2(config.client_id, config.client_secret, config.redirect_uris?.[0]);
auth.setCredentials(tokens);
auth.on('tokens', t => {{
  const c = JSON.parse(fs.readFileSync(path.join(credDir, 'credentials.json'), 'utf-8'));
  Object.assign(c, t);
  fs.writeFileSync(path.join(credDir, 'credentials.json'), JSON.stringify(c, null, 2));
}});
const drive = google.drive({{version:'v3',auth}});
async function main() {{
  const res = await drive.files.get({{fileId:'{DRIVE_FILE_ID}',alt:'media'}},{{responseType:'stream'}});
  const writer = fs.createWriteStream('{XLSX_PATH}');
  await new Promise((resolve,reject) => {{
    res.data.pipe(writer);
    writer.on('finish',resolve);
    writer.on('error',reject);
  }});
  console.log('OK');
}}
main().catch(e=>console.error(e.message));
"""
    script_path = f"/tmp/_drive_dl_{os.getpid()}.js"
    with open(script_path, "w") as f:
        f.write(script)
    
    result = subprocess.run(
        ["node", script_path],
        capture_output=True, text=True, timeout=60,
        env={**os.environ, "HOME": "/home/david",
             "NODE_PATH": f"{NANOCLAW}/node_modules"},
        cwd=NANOCLAW,
    )
    os.unlink(script_path)
    return "OK" in result.stdout


def ensure_tables(conn):
    """Create production tables if they don't exist."""
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS team_production (
        manager TEXT NOT NULL,
        team_name TEXT,
        month TEXT NOT NULL,
        production REAL,
        PRIMARY KEY (manager, month)
    );
    
    CREATE TABLE IF NOT EXISTS agent_production (
        agent_name TEXT NOT NULL,
        team_name TEXT,
        month TEXT NOT NULL,
        production REAL,
        PRIMARY KEY (agent_name, month)
    );
    
    CREATE TABLE IF NOT EXISTS monthly_rankings (
        month TEXT NOT NULL,
        rank INTEGER,
        agent_name TEXT NOT NULL,
        upline TEXT,
        production REAL,
        PRIMARY KEY (month, rank)
    );
    
    CREATE TABLE IF NOT EXISTS screen_names (
        real_name TEXT PRIMARY KEY,
        screen_name TEXT,
        manager TEXT,
        team_name TEXT
    );
    
    CREATE TABLE IF NOT EXISTS monthly_totals (
        month TEXT PRIMARY KEY,
        total_production REAL,
        producing_agents INTEGER,
        pct_change REAL
    );

    CREATE TABLE IF NOT EXISTS sync_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT NOT NULL,
        last_synced TEXT,
        rows_synced INTEGER,
        status TEXT,
        created_at TEXT
    );
    """)


def parse_team_production(wb, conn):
    """Parse the Recruiting Program TEAM Product tab."""
    ws = wb["Recruiting Program TEAM Product"]
    
    # Row 1 is headers: AGENT, TEAM NAMES, AGENT, then monthly dates
    headers = [c.value for c in ws[1]]
    months = []
    for h in headers[3:]:
        if h:
            h_str = str(h)
            if "00:00:00" in h_str:
                # Parse datetime
                try:
                    dt = datetime.strptime(h_str.split(" ")[0], "%Y-%m-%d")
                    months.append(dt.strftime("%Y-%m"))
                except:
                    months.append(h_str[:7])
            else:
                months.append(h_str[:10])
        else:
            months.append(None)
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        manager = row[0]
        team = row[1]
        if not manager or str(manager).startswith("#"):
            continue
        
        for i, val in enumerate(row[3:]):
            if i < len(months) and months[i] and val and str(val) != "#N/A":
                try:
                    prod = float(val)
                    if prod > 0:
                        conn.execute("""
                            INSERT OR REPLACE INTO team_production 
                            (manager, team_name, month, production)
                            VALUES (?, ?, ?, ?)
                        """, (str(manager).strip(), str(team).strip() if team else "",
                              months[i], prod))
                        count += 1
                except (ValueError, TypeError):
                    pass
    
    conn.commit()
    return count


def parse_personal_production(wb, conn):
    """Parse the Recruiting Program Personal Pro tab."""
    ws = wb["Recruiting Program Personal Pro"]
    
    headers = [c.value for c in ws[1]]
    months = []
    for h in headers[2:]:
        if h:
            h_str = str(h)
            if "00:00:00" in h_str:
                try:
                    dt = datetime.strptime(h_str.split(" ")[0], "%Y-%m-%d")
                    months.append(dt.strftime("%Y-%m"))
                except:
                    months.append(h_str[:7])
            else:
                months.append(h_str[:10])
        else:
            months.append(None)
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        team = row[0]
        agent = row[1]
        if not agent or str(agent).startswith("#"):
            continue
        
        for i, val in enumerate(row[2:]):
            if i < len(months) and months[i] and val and str(val) != "#N/A":
                try:
                    prod = float(val)
                    if prod > 0:
                        conn.execute("""
                            INSERT OR REPLACE INTO agent_production
                            (agent_name, team_name, month, production)
                            VALUES (?, ?, ?, ?)
                        """, (str(agent).strip(), str(team).strip() if team else "",
                              months[i], prod))
                        count += 1
                except (ValueError, TypeError):
                    pass
    
    conn.commit()
    return count


def parse_screen_names(wb, conn):
    """Parse the SCREEN NAME tab — nicknames only, NOT hierarchy."""
    ws = wb["SCREEN NAME"]
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        screen = row[1]
        
        if name and str(name).strip():
            conn.execute("""
                INSERT OR REPLACE INTO screen_names
                (real_name, screen_name, manager, team_name)
                VALUES (?, ?, ?, ?)
            """, (str(name).strip(),
                  str(screen).strip() if screen else None,
                  None,  # Do NOT pull manager from this tab
                  None))  # Do NOT pull team from this tab
            count += 1
    
    conn.commit()
    return count


def parse_monthly_totals(wb, conn):
    """Extract monthly totals from the MONTHLY TEAM PRODUCTION tab or summary."""
    # Try the filterable tab first
    count = 0
    for tab_name in wb.sheetnames:
        if "MONTHLY TEAM PRODUCTION FILTERE" in tab_name:
            ws = wb[tab_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Try to find rows with month + total pattern
                vals = [c for c in row if c]
                if len(vals) >= 2:
                    # Look for recognizable month-production pairs
                    pass
            break
    
    # Fallback: use the data we already parsed
    rows = conn.execute("""
        SELECT month, SUM(production) as total, COUNT(DISTINCT agent_name) as agents
        FROM agent_production
        GROUP BY month ORDER BY month
    """).fetchall()
    
    for i, row in enumerate(rows):
        pct = 0
        if i > 0 and rows[i-1][1] > 0:
            pct = round((row[1] - rows[i-1][1]) / rows[i-1][1] * 100, 1)
        conn.execute("""
            INSERT OR REPLACE INTO monthly_totals
            (month, total_production, producing_agents, pct_change)
            VALUES (?, ?, ?, ?)
        """, (row[0], round(row[1], 2), row[2], pct))
        count += 1
    
    conn.commit()
    return count


def main():
    print("=" * 60)
    print("GINA'S PRODUCTION REPORT SYNC")
    print("=" * 60)
    
    print("\n1. Downloading latest from Google Drive...")
    if download_latest():
        size = os.path.getsize(XLSX_PATH)
        print(f"   Downloaded: {size:,} bytes")
    else:
        if os.path.exists(XLSX_PATH):
            print(f"   Using cached file")
        else:
            print("   FAILED — no file available")
            return
    
    print("\n2. Loading workbook...")
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    print(f"   {len(wb.sheetnames)} tabs loaded")
    
    print("\n3. Setting up database...")
    conn = sqlite3.connect(INTEL_DB)
    ensure_tables(conn)
    
    print("\n4. Parsing team production...")
    team_count = parse_team_production(wb, conn)
    print(f"   {team_count} team-month records")
    
    print("\n5. Parsing personal production...")
    agent_count = parse_personal_production(wb, conn)
    print(f"   {agent_count} agent-month records")
    
    print("\n6. Parsing screen names...")
    screen_count = parse_screen_names(wb, conn)
    print(f"   {screen_count} agent screen names")
    
    print("\n7. Building monthly totals...")
    total_count = parse_monthly_totals(wb, conn)
    print(f"   {total_count} months")
    
    # Summary
    print("\n" + "=" * 60)
    
    top_teams = conn.execute("""
        SELECT manager, team_name, SUM(production) as total
        FROM team_production GROUP BY manager
        ORDER BY total DESC LIMIT 10
    """).fetchall()
    print("\nTOP TEAMS (all time):")
    for t in top_teams:
        print(f"  {t[0]:25s} {t[1]:30s} ${t[2]:>12,.2f}")
    
    top_agents = conn.execute("""
        SELECT agent_name, team_name, SUM(production) as total
        FROM agent_production GROUP BY agent_name
        ORDER BY total DESC LIMIT 10
    """).fetchall()
    print("\nTOP AGENTS — PERSONAL PRODUCTION (all time):")
    for a in top_agents:
        print(f"  {a[0]:25s} {a[1]:30s} ${a[2]:>12,.2f}")
    
    latest = conn.execute("""
        SELECT month, total_production, producing_agents, pct_change
        FROM monthly_totals ORDER BY month DESC LIMIT 6
    """).fetchall()
    print("\nRECENT MONTHLY TOTALS:")
    for m in reversed(latest):
        print(f"  {m[0]}  ${m[1]:>12,.2f}  {m[2]} agents  ({m[3]:+.1f}%)")
    
    print("=" * 60)
    
    # Log sync
    conn.execute("""
        INSERT INTO sync_log (source, last_synced, rows_synced, status, created_at)
        VALUES (?, ?, ?, ?, datetime('now'))
    """, ("gina_production", datetime.utcnow().isoformat(),
          team_count + agent_count + screen_count, "success"))
    conn.commit()
    conn.close()


if __name__ == "__main__":
    main()
