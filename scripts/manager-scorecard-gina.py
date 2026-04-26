#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('/home/david/nanoclaw/data/production-reporting.xlsx', data_only=True)
ws = wb['Recruiting Program TEAM Product']

print('MANAGER SCORECARD — FROM GINA\'S PRODUCTION REPORT')
print('=' * 90)
print()

# Get month headers
headers = []
for col in range(3, min(ws.max_column+1, 20)):
    val = ws.cell(1, col).value
    if val:
        h = str(val)[:10]
        headers.append(h)

# Print each manager with their team production
for row_idx in range(2, ws.max_row+1):
    manager = ws.cell(row_idx, 1).value
    team = ws.cell(row_idx, 2).value
    agent_col = ws.cell(row_idx, 3).value
    if not manager:
        continue
    
    manager = str(manager).strip()
    team = str(team).strip() if team else ''
    
    # Get monthly values
    monthly = {}
    total = 0
    for col_idx in range(3, min(ws.max_column+1, 20)):
        val = ws.cell(row_idx, col_idx).value
        if val and str(val) != '#N/A':
            try:
                amt = float(val)
                if amt > 0:
                    total += amt
                    month_header = headers[col_idx-3] if col_idx-3 < len(headers) else '?'
                    monthly[month_header] = amt
            except (ValueError, TypeError):
                pass
    
    if total > 0 or team:
        months_str = len(monthly)
        latest = list(monthly.values())[-1] if monthly else 0
        print(f'  {manager:25s} {team:35s}')
        print(f'    Total: ${total:>12,.2f} | Months: {months_str} | Latest month: ${latest:>10,.2f}')
        print()
